import openpyxl


def get_test_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    data = []

    for col in range(1, sheet.max_column + 1):
        data.append(sheet.cell(row=2, column=col).value)

    return data